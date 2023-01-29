import baostock as bs
import pandas as pd
import talib as ta
import datetime
import os
import multiprocessing
import yaml
import numpy as np
import openpyxl as op


def get_config():
    with open("./config.yaml", encoding='utf-8') as f:
        config = yaml.safe_load(f)
        return config


def get_gold():
    bs.login()
    starttime = datetime.datetime.now()

    cfg = get_config()

    end_date = (lambda x: datetime.datetime.strptime(x, '%Y-%m-%d').date() if x != '' else datetime.date.today())(
        cfg['end_date'])
    end_date_str = end_date.strftime('%Y-%m-%d')
    start_date = end_date - datetime.timedelta(days=2 * cfg['ma']['days'])
    start_date_str = start_date.strftime('%Y-%m-%d')

    stock_rs = bs.query_all_stock(end_date.strftime('%Y-%m-%d'))
    if stock_rs.error_code != '0' or len(stock_rs.data) == 0:
        return bs.logout()
    stock_df = stock_rs.get_data()

    pool = multiprocessing.Pool(processes=os.cpu_count())
    macd_list = []
    kdj_list = []
    ma250_list = []
    for code in stock_df["code"]:
        print("Computing :" + code)
        resp_data = get_history_k_data(bs, code, start_date_str, end_date_str)
        if resp_data.error_code != '0' or len(resp_data.data) == 0:
            continue
        df = pd.DataFrame(resp_data.data, columns=resp_data.fields)

        macd_list.append(pool.apply_async(compute_macd, args=(df, code, end_date, cfg)))
        kdj_list.append(pool.apply_async(compute_kdj, args=(df, code, end_date, cfg)))
        ma250_list.append(pool.apply_async(compute_ma250, args=(df, code, cfg)))

    macd_code = [l.get() for l in macd_list if l.get() is not None]
    kdj_code = [l.get() for l in kdj_list if l.get() is not None]
    ma250_code = [l.get() for l in ma250_list if l.get() is not None]
    print('MACD金叉 + MA250结果:', set(macd_code) & set(ma250_code))
    print('MACD金叉 + KDJ金叉 + MA250结果:', set(macd_code) & set(kdj_code) & set(ma250_code))
    write_to_xlsx(end_date_str, '选股结果.xlsx', 'MACD+MA250', list(set(macd_code) & set(ma250_code)))
    write_to_xlsx(end_date_str, '选股结果.xlsx', 'MACD+KDJ+MA250',
                  list(set(macd_code) & set(kdj_code) & set(ma250_code)))
    bs.logout()
    endtime = datetime.datetime.now()
    print('程序执行时长(s):', (endtime - starttime).seconds)


def write_to_xlsx(date_str, filename, sheetname, data):
    if not os.path.exists('./' + date_str):
        os.makedirs('./' + date_str)
    wb = op.Workbook()
    filepath = os.path.join(date_str, filename)
    if os.path.exists(filepath):
        wb = op.load_workbook(filepath)
    wb.create_sheet(title=sheetname)
    ws = wb[sheetname]
    for i, d in enumerate(data, 1):
        ws.cell(i, 1, value=d)
        ws.cell(i, 1).hyperlink = 'http://quote.eastmoney.com/' + d.replace('.', '') + '.html#fullScreenChart'
        ws.cell(i, 1).style = "Hyperlink"
    wb.save(os.path.join(date_str, filename))


def get_history_k_data(bsclient, code, startdate, enddate):
    # 获取股票日 K 线数据
    return bsclient.query_history_k_data_plus(code,
                                              "date,code,high,low,close,tradestatus",
                                              start_date=startdate,
                                              end_date=enddate,
                                              frequency="d", adjustflag="3")


def compute_macd(df, code, end_date, cfg):
    # 剔除停盘数据
    df2 = df[df['tradestatus'] == '1']
    # 获取 dif,dea,hist，它们的数据类似是 tuple，且跟 df2 的 date 日期一一对应
    # 记住了 dif,dea,hist 前 33 个为 Nan，所以推荐用于计算的数据量一般为你所求日期之间数据量的3倍
    # 这里计算的 hist 就是 dif-dea,而很多证券商计算的 MACD=hist*2=(difdea)*2
    dif, dea, hist = ta.MACD(df2['close'].astype(float).values, fastperiod=10, slowperiod=22, signalperiod=7)
    df_data = pd.DataFrame({'dif': dif[33:], 'dea': dea[33:], 'hist': hist[33:]}, index=df2['date'][33:],
                           columns=['dif', 'dea', 'hist'])
    # df_data.plot(title='MACD')
    # 寻找 MACD 金叉和死叉
    macd_position = df_data['dif'] > df_data['dea']
    cross_date = macd_position[(macd_position == True) & (macd_position.shift() == False)].index.values
    try:
        last_date = datetime.datetime.strptime(cross_date[-1], '%Y-%m-%d').date()
        if 0 <= (end_date - last_date).days <= cfg['macd_cross_days']:
            return code
    except:
        return


def compute_kdj(df, code, end_date, cfg):
    # 剔除停盘数据
    df_status = df[df['tradestatus'] == '1']
    # 计算KDJ指标,前9个数据为空
    low_list = df_status['low'].rolling(window=9).min()
    high_list = df_status['high'].rolling(window=9).max()
    rsv = (df_status['close'].astype(float) - low_list) / (high_list - low_list) * 100
    df_data = pd.DataFrame()
    df_data['K'] = rsv.ewm(com=2).mean()
    df_data['D'] = df_data['K'].ewm(com=2).mean()
    df_data['J'] = 3 * df_data['K'] - 2 * df_data['D']
    df_data.index = df_status['date'].values
    df_data.index.name = 'date'
    # 删除空数据
    df_data = df_data.dropna()
    # 计算KDJ指标金叉、死叉情况
    kdj_position = df_data['K'] > df_data['D']
    # df_data.plot(title="KDJ")

    cross_date = kdj_position[(kdj_position == True) & (kdj_position.shift() == False)].index.values
    try:
        last_date = datetime.datetime.strptime(cross_date[-1], '%Y-%m-%d').date()
        if 0 <= (end_date - last_date).days <= cfg['kdj_cross_days']:
            return code
    except:
        return


def compute_ma250(df, code, cfg):
    df['MA250'] = ta.MA(df['close'].astype(float).values, timeperiod=cfg['ma']['days'], matype=0)
    count = 0
    percent = cfg['ma']['percent']
    for i, ma in df[-5:].iterrows():
        if ma['MA250'] * (1 - percent) <= float(ma['close']) <= ma['MA250'] * (1 + percent):
            count += 1
        if count == 3 and (
                trend_line(df[-5:].index.values, df[-5:]['MA250'].values) >= 0 if cfg['ma'][
                    'trend_raise'] else True):
            return code


def trend_line(index, data, order=1):
    coeffs = np.polyfit(index, list(data), order)
    slope = coeffs[-2]
    return slope


if __name__ == '__main__':
    get_gold()
